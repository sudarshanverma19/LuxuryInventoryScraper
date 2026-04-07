"""
Export service — generates CSV and Excel files from product data.
"""

import io
import csv
import logging
from typing import Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from database.models import Brand, Product, Variant

logger = logging.getLogger(__name__)


async def export_products(
    session: AsyncSession,
    format: str = "csv",
    brand_slug: Optional[str] = None,
) -> tuple[bytes, str, str]:
    """
    Export products with variants to CSV or Excel.
    Returns (file_bytes, filename, content_type).
    """
    # Fetch data
    query = (
        select(Product, Variant, Brand)
        .join(Variant, Variant.product_id == Product.id, isouter=True)
        .join(Brand, Product.brand_id == Brand.id)
    )

    if brand_slug:
        query = query.where(Brand.slug == brand_slug)

    query = query.order_by(Brand.name, Product.name, Variant.color, Variant.size)

    result = await session.execute(query)
    rows = result.all()

    # Build flat rows
    headers = [
        "Brand", "Product Name", "Category", "Price", "Currency",
        "Color", "Size", "In Stock", "Quantity", "SKU",
        "Product URL", "Image URL",
    ]

    data_rows = []
    for product, variant, brand in rows:
        row = [
            brand.name,
            product.name,
            product.category or "",
            product.price if product.price else "",
            product.currency or "USD",
            variant.color if variant else "",
            variant.size if variant else "",
            "Yes" if (variant and variant.in_stock) else "No",
            variant.quantity if (variant and variant.quantity is not None) else "",
            variant.sku if variant else "",
            product.url,
            product.image_url or "",
        ]
        data_rows.append(row)

    if format == "xlsx":
        return _to_excel(headers, data_rows, brand_slug)
    else:
        return _to_csv(headers, data_rows, brand_slug)


def _to_csv(headers, data_rows, brand_slug) -> tuple[bytes, str, str]:
    """Generate CSV file."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(data_rows)

    filename = f"inventory_{brand_slug or 'all'}.csv"
    return output.getvalue().encode("utf-8"), filename, "text/csv"


def _to_excel(headers, data_rows, brand_slug) -> tuple[bytes, str, str]:
    """Generate Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Style header row
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7C5CFC", end_color="7C5CFC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # Color-code stock status
            if headers[col_idx - 1] == "In Stock":
                if value == "Yes":
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row in data_rows:
            val = str(row[col_idx - 1])
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = min(max_len + 4, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)

    filename = f"inventory_{brand_slug or 'all'}.xlsx"
    return output.getvalue(), filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
